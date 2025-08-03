from django.contrib import admin
from django.db.models import Count
from .models import User, Poll, Question, Choice, Vote
import csv
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter

# --- Export Actions for Questions ---

def export_votes_csv(modeladmin, request, queryset):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="votes.csv"'

    writer = csv.writer(response)
    writer.writerow(['Poll', 'Question', 'Choice', 'Vote Count'])

    for question in queryset:
        for choice in question.choices.annotate(vote_count=Count('votes')):
            writer.writerow([
                question.poll.title,
                question.text,
                choice.text,
                choice.vote_count
            ])
    return response
export_votes_csv.short_description = "Export Votes as CSV"

def export_votes_excel(modeladmin, request, queryset):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Vote Stats"

    headers = ['Poll', 'Question', 'Choice', 'Vote Count']
    sheet.append(headers)

    for question in queryset:
        for choice in question.choices.annotate(vote_count=Count('votes')):
            sheet.append([
                question.poll.title,
                question.text,
                choice.text,
                choice.vote_count
            ])

    for i, column in enumerate(sheet.columns, start=1):
        max_length = max(len(str(cell.value or "")) for cell in column)
        sheet.column_dimensions[get_column_letter(i)].width = max_length + 2

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="votes.xlsx"'
    workbook.save(response)
    return response
export_votes_excel.short_description = "Export Votes as Excel"

# --- Export Actions for Polls with Percentages ---

def export_poll_votes_csv(modeladmin, request, queryset):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="poll_vote_stats.csv"'

    writer = csv.writer(response)
    writer.writerow(['Poll', 'Question', 'Choice', 'Votes', 'Percentage'])

    for poll in queryset:
        questions = poll.questions.all()
        for question in questions:
            choices = question.choices.annotate(vote_count=Count('votes'))
            total_votes = sum(choice.vote_count for choice in choices)
            for choice in choices:
                percent = (choice.vote_count / total_votes * 100) if total_votes > 0 else 0
                writer.writerow([
                    poll.title,
                    question.text,
                    choice.text,
                    choice.vote_count,
                    f"{percent:.2f}%"
                ])

    return response
export_poll_votes_csv.short_description = "Export Poll Votes as CSV"

def export_poll_votes_excel(modeladmin, request, queryset):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Poll Vote Stats"

    headers = ['Poll', 'Question', 'Choice', 'Votes', 'Percentage']
    sheet.append(headers)

    for poll in queryset:
        questions = poll.questions.all()
        for question in questions:
            choices = question.choices.annotate(vote_count=Count('votes'))
            total_votes = sum(choice.vote_count for choice in choices)
            for choice in choices:
                percent = (choice.vote_count / total_votes * 100) if total_votes > 0 else 0
                sheet.append([
                    poll.title,
                    question.text,
                    choice.text,
                    choice.vote_count,
                    f"{percent:.2f}%"
                ])

    for i, column in enumerate(sheet.columns, start=1):
        max_length = max(len(str(cell.value or "")) for cell in column)
        sheet.column_dimensions[get_column_letter(i)].width = max_length + 2

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="poll_vote_stats.xlsx"'
    workbook.save(response)
    return response
export_poll_votes_excel.short_description = "Export Poll Votes as Excel"

# --- Admin Registrations ---

@admin.register(User)
class UserAdmin(admin.ModelAdmin):
    list_display = ('id', 'username', 'email', 'date_joined', 'is_active')

@admin.register(Poll)
class PollAdmin(admin.ModelAdmin):
    list_display = ('id', 'title', 'created_at', 'total_votes')
    actions = [export_poll_votes_csv, export_poll_votes_excel]

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        return qs.annotate(vote_count=Count('questions__choices__votes'))

    def total_votes(self, obj):
        return obj.vote_count
    total_votes.short_description = 'Total Votes'

@admin.register(Question)
class QuestionAdmin(admin.ModelAdmin):
    list_display = ('id', 'poll', 'text', 'vote_count', 'get_winner')
    list_filter = ('poll',)
    actions = [export_votes_csv, export_votes_excel]

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        return qs.annotate(vote_count=Count('choices__votes')).order_by('-vote_count')

    def vote_count(self, obj):
        return obj.vote_count
    vote_count.admin_order_field = 'vote_count'
    vote_count.short_description = 'Total Votes'

    def get_winner(self, obj):
        winner = (
            obj.choices.annotate(vote_count=Count('votes'))
            .order_by('-vote_count')
            .first()
        )
        if winner and winner.vote_count > 0:
            return f"{winner.text} ({winner.vote_count} votes)"
        return "No votes yet"
    get_winner.short_description = "Winning Choice"

@admin.register(Choice)
class ChoiceAdmin(admin.ModelAdmin):
    list_display = ('id', 'question', 'text')

@admin.register(Vote)
class VoteAdmin(admin.ModelAdmin):
    list_display = ('id', 'choice', 'user', 'voted_at')
